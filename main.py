from tkinter import *
from tkinter.ttk import Combobox
from tkinter import messagebox
from openpyxl import Workbook
import openpyxl , pathlib

file = pathlib.Path('data/Backend_data.xlsx')

if file.exists():
    pass
else:
    file = Workbook()
    sheet = file.active
    sheet['A1'] = 'Full Name'
    sheet['B1'] = 'Phone Number'
    sheet['C1'] = 'Age'
    sheet['D1'] = 'Gender'
    sheet['E1'] = 'Address'
    file.save('data/Backend_data.xlsx')


def submit():
    name = nameVar.get()
    contactno = contactnoVar.get()
    age = ageVar.get()
    gender = genderCombo.get()
    address = addressText.get(1.0,END)

    file = openpyxl.load_workbook('data/Backend_data.xlsx')
    sheet= file.active

    sheet.cell(column=1, row=sheet.max_row+1, value=name)
    sheet.cell(column=2, row=sheet.max_row, value=contactno)
    sheet.cell(column=3, row=sheet.max_row, value=age)
    sheet.cell(column=4, row=sheet.max_row, value=gender)
    sheet.cell(column=5, row=sheet.max_row, value=address)

    file.save('data/Backend_data.xlsx')
    messagebox.showinfo('Info', 'Details added!')
    clear()



def clear():
    nameVar.set('')
    contactnoVar.set('')
    ageVar.set('')
    genderCombo.set('Male')
    addressText.delete(1.0,END)



if __name__ == '__main__':
    root = Tk()
    root.title("Students Management System")
    root.geometry('900x540+500+300')
    root.resizable(False, False)
    root.config(bg='black')

    Label(root, text="Students Management System", font=('Berlin Sans FB', '45'), bg='black', fg='darkorange').pack(pady=10)

    Label(root, text='Fill out this form to register.', font=('Constantia', '17'), bg='black', fg='white').place(x=305, y=80)
    Label(root, text='Name:', font=('Constantia', '22'), bg='black', fg='white').place(x=120, y=120)
    Label(root, text='Contact No:', font=('Constantia', '22'), bg='black', fg='white').place(x=120, y=180)
    Label(root, text='Age:', font=('Constantia', '22'), bg='black', fg='white').place(x=120, y=240)
    Label(root, text='Gender:', font=('Constantia', '22'), bg='black', fg='white').place(x=530, y=240)
    Label(root, text='Address:', font=('Constantia', '22'), bg='black', fg='white').place(x=120, y=300)

    nameVar = StringVar()
    contactnoVar = StringVar()
    ageVar = StringVar()

    Entry(root, font=('Constantia', '18'), bd=2, relief='groove', textvariable=nameVar).place(x=300, y=125, width=550, height=35)
    Entry(root, font=('Constantia', '18'), bd=2, relief='groove', textvariable=contactnoVar).place(x=300, y=185, width=550, height=35)
    Entry(root, font=('Constantia', '18'), bd=2, relief='groove', textvariable=ageVar).place(x=300, y=245, width=210, height=35)

    genderCombo = Combobox(root, font=('Constantia', '18'), values=['Male', 'Female'], state='r')
    genderCombo.place(x=660, y=245, width=190, height=35)
    genderCombo.set('Male')

    addressText = Text(root, font=('Constantia', '20'), bd=2, relief='groove')
    addressText.place(x=300, y=300, width=550, height=150)

    Button(root,text='Submit',font=('Berlin Sans Fb', '22'), bg='black', fg='white', command=submit).place(x=320, y=465, width=160)
    Button(root,text='Clear',font=('Berlin Sans Fb', '22'), bg='black', fg='white', command=clear).place(x=500, y=465, width=160)
    Button(root,text='Exit',font=('Berlin Sans Fb', '22'), bg='black', fg='white', command=root.destroy).place(x=680, y=465, width=160)

    root.mainloop()